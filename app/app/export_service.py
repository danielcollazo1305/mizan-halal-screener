"""
export_service.py
-----------------
Generates PDF and Excel exports for Mizan Halal Screener.
"""
import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── PDF Export ────────────────────────────────────────────────────────────────

def generate_stock_pdf(data: dict) -> bytes:
    """Generates a PDF report for a single stock analysis."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=20*mm, leftMargin=20*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )

    # Colors
    GREEN  = colors.HexColor("#22c55e")
    DARK   = colors.HexColor("#0f172a")
    SLATE  = colors.HexColor("#1e293b")
    GRAY   = colors.HexColor("#94a3b8")
    RED    = colors.HexColor("#ef4444")
    YELLOW = colors.HexColor("#f59e0b")

    status_color = {"HALAL": GREEN, "QUESTIONABLE": YELLOW, "HARAM": RED}
    color = status_color.get(data.get("status", ""), GRAY)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontSize=22, textColor=GREEN, spaceAfter=4, fontName="Helvetica-Bold")
    sub_style   = ParagraphStyle("sub",   fontSize=11, textColor=GRAY,  spaceAfter=2, fontName="Helvetica")
    label_style = ParagraphStyle("label", fontSize=9,  textColor=GRAY,  fontName="Helvetica")
    body_style  = ParagraphStyle("body",  fontSize=10, textColor=colors.HexColor("#f1f5f9"), fontName="Helvetica", leading=14)

    story = []

    # Header
    story.append(Paragraph("🕌 Mizan Halal Screener", title_style))
    story.append(Paragraph("Stock Analysis Report", sub_style))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", label_style))
    story.append(HRFlowable(width="100%", thickness=1, color=GREEN, spaceAfter=12))

    # Stock header
    story.append(Paragraph(f"{data.get('name', '')} ({data.get('ticker', '')})", ParagraphStyle("h1", fontSize=18, fontName="Helvetica-Bold", textColor=colors.HexColor("#f1f5f9"), spaceAfter=4)))
    story.append(Paragraph(f"{data.get('sector', '')} · {data.get('industry', '')} · {data.get('country', '')}", sub_style))
    story.append(Spacer(1, 8))

    # Status + Score table
    status = data.get("status", "N/A")
    price  = data.get("price")
    score  = data.get("investment_score") or data.get("fundamental_score", "N/A")
    grade  = data.get("grade", "N/A")

    summary_data = [
        ["Price", "Status", "Grade", "Investment Score"],
        [
            f"${price:.2f}" if price else "N/A",
            status,
            str(grade),
            str(score),
        ]
    ]
    summary_table = Table(summary_data, colWidths=[40*mm, 45*mm, 35*mm, 45*mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), SLATE),
        ("TEXTCOLOR",   (0,0), (-1,0), GRAY),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,0), 9),
        ("BACKGROUND",  (0,1), (-1,1), DARK),
        ("TEXTCOLOR",   (0,1), (-1,1), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR",   (1,1), (1,1),  color),
        ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,1), (-1,1), 13),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [SLATE, DARK]),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#334155")),
        ("TOPPADDING",  (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0),(-1,-1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 16))

    # Halal reason
    reason = data.get("reason", "")
    if reason:
        story.append(Paragraph("Halal Status", ParagraphStyle("h2", fontSize=12, fontName="Helvetica-Bold", textColor=GREEN, spaceAfter=6)))
        story.append(Paragraph(reason, body_style))
        story.append(Spacer(1, 12))

    # Fundamentals
    story.append(Paragraph("Key Fundamentals", ParagraphStyle("h2", fontSize=12, fontName="Helvetica-Bold", textColor=GREEN, spaceAfter=6)))

    def pct(v): return f"{v*100:.1f}%" if v is not None else "N/A"
    def fmt(v, prefix="$"): return f"{prefix}{v:.2f}" if v is not None else "N/A"

    fundamentals = [
        ["Metric", "Value", "Metric", "Value"],
        ["Price",           fmt(data.get("price")),              "Market Cap",      _fmt_large(data.get("market_cap"))],
        ["ROE",             pct(data.get("roe")),                "Profit Margin",   pct(data.get("profit_margin"))],
        ["P/E Ratio",       fmt(data.get("pe_ratio"), ""),       "P/B Ratio",       fmt(data.get("pb_ratio"), "")],
        ["Revenue Growth",  pct(data.get("revenue_growth")),     "Earnings Growth", pct(data.get("earnings_growth"))],
        ["Debt Ratio",      pct(data.get("debt_ratio")),         "Dividend Yield",  fmt(data.get("dividend_yield"), "") + "%" if data.get("dividend_yield") else "N/A"],
        ["EPS",             fmt(data.get("eps")),                "Book Value",      fmt(data.get("book_value"))],
    ]

    fund_table = Table(fundamentals, colWidths=[45*mm, 40*mm, 45*mm, 40*mm])
    fund_table.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), SLATE),
        ("TEXTCOLOR",   (0,0), (-1,0), GRAY),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("BACKGROUND",  (0,1), (0,-1), colors.HexColor("#1a2744")),
        ("BACKGROUND",  (2,1), (2,-1), colors.HexColor("#1a2744")),
        ("TEXTCOLOR",   (0,1), (-1,-1), colors.HexColor("#f1f5f9")),
        ("TEXTCOLOR",   (1,1), (1,-1), GREEN),
        ("TEXTCOLOR",   (3,1), (3,-1), GREEN),
        ("FONTNAME",    (1,1), (1,-1), "Helvetica-Bold"),
        ("FONTNAME",    (3,1), (3,-1), "Helvetica-Bold"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#334155")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [DARK, SLATE]),
        ("TOPPADDING",  (0,0), (-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("ALIGN",       (1,0), (1,-1), "CENTER"),
        ("ALIGN",       (3,0), (3,-1), "CENTER"),
    ]))
    story.append(fund_table)
    story.append(Spacer(1, 16))

    # Fair Value
    fv = data.get("fair_value", {})
    if fv:
        story.append(Paragraph("Fair Value Analysis", ParagraphStyle("h2", fontSize=12, fontName="Helvetica-Bold", textColor=GREEN, spaceAfter=6)))
        fv_data = [
            ["Graham Value", "DCF Value", "Upside", "Valuation"],
            [
                f"${fv.get('graham_value', 0):.2f}" if fv.get('graham_value') else "N/A",
                f"${fv.get('dcf_value', 0):.2f}"    if fv.get('dcf_value')    else "N/A",
                f"{fv.get('upside_pct', 0):+.1f}%"  if fv.get('upside_pct') is not None else "N/A",
                str(fv.get("valuation", "N/A")),
            ]
        ]
        fv_table = Table(fv_data, colWidths=[42*mm, 42*mm, 42*mm, 42*mm])
        fv_table.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), SLATE),
            ("TEXTCOLOR",   (0,0), (-1,0), GRAY),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("BACKGROUND",  (0,1), (-1,1), DARK),
            ("TEXTCOLOR",   (0,1), (-1,1), GREEN),
            ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 10),
            ("ALIGN",       (0,0), (-1,-1), "CENTER"),
            ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#334155")),
            ("TOPPADDING",  (0,0), (-1,-1), 8),
            ("BOTTOMPADDING",(0,0),(-1,-1), 8),
        ]))
        story.append(fv_table)
        story.append(Spacer(1, 16))

    # Footer
    story.append(HRFlowable(width="100%", thickness=1, color=GRAY, spaceBefore=8))
    story.append(Paragraph(
        "⚠️ This report is for informational purposes only. Not financial advice. Always do your own research.",
        ParagraphStyle("footer", fontSize=8, textColor=GRAY, alignment=TA_CENTER)
    ))
    story.append(Paragraph("mizan-web-omega.vercel.app", ParagraphStyle("link", fontSize=8, textColor=GREEN, alignment=TA_CENTER)))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _fmt_large(v):
    if not v: return "N/A"
    if v >= 1e12: return f"${v/1e12:.2f}T"
    if v >= 1e9:  return f"${v/1e9:.2f}B"
    if v >= 1e6:  return f"${v/1e6:.2f}M"
    return f"${v:.2f}"


# ── Excel Export ──────────────────────────────────────────────────────────────

def generate_ranking_excel(companies: list[dict]) -> bytes:
    """Generates an Excel file with the full ranking."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mizan Halal Ranking"

    # Colors
    GREEN_FILL  = PatternFill("solid", fgColor="22c55e")
    DARK_FILL   = PatternFill("solid", fgColor="0f172a")
    SLATE_FILL  = PatternFill("solid", fgColor="1e293b")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    GREEN_FONT  = Font(bold=True, color="22c55e", size=11)

    # Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "🕌 Mizan Halal Screener — Stock Ranking"
    ws["A1"].font = Font(bold=True, color="22c55e", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = DARK_FILL

    ws.merge_cells("A2:N2")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font = Font(color="94a3b8", size=9)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].fill = DARK_FILL

    # Headers
    headers = [
        "Ticker", "Name", "Status", "Grade", "Score",
        "Price", "Market Cap", "P/E", "P/B",
        "ROE", "Profit Margin", "Revenue Growth",
        "Debt Ratio", "Upside %"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = SLATE_FILL
        cell.alignment = Alignment(horizontal="center")

    # Data
    def pct(v): return f"{v*100:.1f}%" if v is not None else "N/A"
    def fmt_large(v):
        if not v: return "N/A"
        if v >= 1e12: return f"${v/1e12:.2f}T"
        if v >= 1e9:  return f"${v/1e9:.2f}B"
        if v >= 1e6:  return f"${v/1e6:.2f}M"
        return f"${v:.2f}"

    status_colors = {"HALAL": "22c55e", "QUESTIONABLE": "f59e0b", "HARAM": "ef4444"}

    for row_idx, c in enumerate(companies, 5):
        status = c.get("status", "")
        s_color = status_colors.get(status, "94a3b8")
        fv = c.get("fair_value", {})

        row_data = [
            c.get("ticker", ""),
            c.get("name", ""),
            status,
            c.get("grade", ""),
            c.get("investment_score") or c.get("fundamental_score", ""),
            f"${c.get('price', 0):.2f}" if c.get("price") else "N/A",
            fmt_large(c.get("market_cap")),
            f"{c.get('pe_ratio', 0):.1f}" if c.get("pe_ratio") else "N/A",
            f"{c.get('pb_ratio', 0):.1f}" if c.get("pb_ratio") else "N/A",
            pct(c.get("roe")),
            pct(c.get("profit_margin")),
            pct(c.get("revenue_growth")),
            pct(c.get("debt_ratio")),
            f"{fv.get('upside_pct', 0):+.1f}%" if fv.get("upside_pct") is not None else "N/A",
        ]

        fill = PatternFill("solid", fgColor="0f172a" if row_idx % 2 == 0 else "1e293b")
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(color="f1f5f9", size=10)
            if col_idx == 3:  # Status column
                cell.font = Font(color=s_color, bold=True, size=10)

    # Column widths
    widths = [8, 28, 14, 8, 8, 10, 12, 8, 8, 10, 14, 16, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Footer
    last_row = len(companies) + 6
    ws.merge_cells(f"A{last_row}:N{last_row}")
    ws[f"A{last_row}"] = "⚠️ Not financial advice. For informational purposes only. | mizan-web-omega.vercel.app"
    ws[f"A{last_row}"].font = Font(color="475569", size=8, italic=True)
    ws[f"A{last_row}"].alignment = Alignment(horizontal="center")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()